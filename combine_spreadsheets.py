from openpyxl import Workbook, load_workbook
import os


def compile_workbooks(workbooks_path, final_filename):
    if not isinstance(workbooks_path, str):
        raise TypeError("Argument workbooks_path must be of type str.")
    
    if not isinstance(final_filename, str):
        raise TypeError("Argument final_filename must be of type str.")

    if not os.path.exists(workbooks_path):
        raise NotADirectoryError("Argument workbook_path is not a directory.")

    if not final_filename.endswith(".xlsx"):
        raise ValueError('final_filename must end with the string ".xlsx"')
        
    if final_filename in os.listdir(workbooks_path):
        raise ValueError(f'There is already a file named {final_filename} in {workbooks_path}. '
                         f'Remove this file first or change the final_filename parameter value.')
    
    wbs = []
    for file in os.listdir(workbooks_path):
        if not file.startswith("~$") and file.endswith(".xlsx"):
            wb = load_workbook(os.path.join(workbooks_path, file))
            wbs.append(wb)
    
    final_wb = Workbook()
    final_ws = final_wb.worksheets[0]

    wb1 = wbs[0]
    ws1 = wb1.worksheets[0] 
 
    for j in range(1, ws1.max_column+1):
        final_ws.cell(row=1, column=j).value = ws1.cell(row=1, column=j).value

    current_row = 2

    for wb in wbs:
        for ws in wb.worksheets:
            mr = ws.max_row 
            mc = ws.max_column 

            for i in range (2, mr + 1): 
                for j in range (1, mc + 1): 
                    current_cell = ws.cell(row = i, column = j) 
                    final_ws.cell(row = current_row, column = j).value = current_cell.value

                current_row += 1

    final_wb.save(os.path.join(workbooks_path, final_filename))


if __name__ == '__main__':
    compile_workbooks(os.path.join(os.getcwd(), "CHANGE_TO_FOLDER_WITH_SPREADSHEETS"), "final.xlsx")
